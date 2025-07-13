from kiteconnect import KiteConnect
import pandas as pd
import sqlite3
from datetime import datetime
from openpyxl import load_workbook
import time
import warnings
import os

warnings.filterwarnings("ignore", message="Could not infer format, so each element will be parsed individually*")

# ====== CONFIG ======
api_key = 'your_api_key_here'
api_secret = 'your_api_secret_here'

access_token_path = os.path.join(os.path.dirname(__file__), 'AccessToken.txt')
db_path = os.path.join(os.path.dirname(__file__), 'Practice.db')
log_path = os.path.join(os.path.dirname(__file__), 'Log.xlsx')

symbols = ['HB', 'RS', 'IB', 'ISS']
product_type = 'MIS'
quantity = 1
poll_interval = 5  # seconds

# ====== INIT KITE ======
kite = KiteConnect(api_key=api_key)
with open(access_token_path) as f:
    access_token = f.read().strip()
kite.set_access_token(access_token)

# ====== UTILITY FUNCTIONS ======

def log_action(message):
    """Log activity as a single string to 'Sheet1' in Excel log file."""
    try:
        book = load_workbook(log_path)
        if 'Sheet1' not in book.sheetnames:
            sheet = book.create_sheet('Sheet1')
        else:
            sheet = book['Sheet1']
            
        if 'error' not in message:
            sheet.append([f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {message}"])
            book.save(log_path)
    except Exception as e:
        print(f"Log write error: {e}")

def log_price(execution_time, price, transaction_type):
    """Log executed price to Sheet2 with + for Buy and - for Sell."""
    try:
        book = load_workbook(log_path)
        if 'Sheet2' not in book.sheetnames:
            sheet = book.create_sheet('Sheet2')
            sheet.append(['Timestamp', 'Executed Price'])
        else:
            sheet = book['Sheet2']
        signed_price = price if transaction_type.upper() == 'BUY' else -price
        sheet.append([execution_time.strftime('%Y-%m-%d %H:%M:%S'), signed_price])
        book.save(log_path)
    except Exception as e:
        print(f"Price log write error: {e}")

def log_trade_taken(symbol, side):
    """Log new open trade in Sheet3."""
    try:
        book = load_workbook(log_path)
        if 'Sheet3' not in book.sheetnames:
            sheet = book.create_sheet('Sheet3')
            sheet.append(['Timestamp', 'Symbol', 'Side'])
        else:
            sheet = book['Sheet3']
        sheet.append([datetime.now().strftime('%Y-%m-%d %H:%M:%S'), symbol, side])
        book.save(log_path)
    except Exception as e:
        log_action(f"Trade log write error: {e}")

def get_live_trade():
    """Check Sheet3 for any open trade."""
    try:
        book = load_workbook(log_path)
        if 'Sheet3' not in book.sheetnames:
            sheet = book.create_sheet('Sheet3')
            book.save(log_path)
            return None
        sheet = book['Sheet3']
        trades = list(sheet.iter_rows(values_only=True))
        if len(trades) < 2:
            return None
        last = trades[-1]
        return {'timestamp': last[0], 'symbol': last[1], 'side': last[2]}
    except Exception as e:
        log_action(f"Live trade check error: {e}")
        return None

def remove_live_trade():
    """Remove last trade from Sheet3."""
    try:
        book = load_workbook(log_path)
        if 'Sheet3' in book.sheetnames:
            sheet = book['Sheet3']
            max_row = sheet.max_row
            if max_row > 1:
                sheet.delete_rows(max_row, 1)
                book.save(log_path)
    except Exception as e:
        log_action(f"Trade log delete error: {e}")

def is_order_pending():
    """Check for any active positions or open orders."""
    try:
        positions = kite.positions()
        net_positions = positions['net']
        for p in net_positions:
            if p['quantity'] != 0:
                return True
        orders = kite.orders()
        for o in orders:
            if o['status'] in ['OPEN', 'TRIGGER PENDING']:
                return True
        return False
    except Exception as e:
        log_action(f"Check pending order error: {e}")
        return True

def map_symbol(symbol):
    """Map short code to trading symbol."""
    mapping = {'HB': 'HDFCBANK', 'RS': 'RELIANCE', 'IB': 'ICICIBANK', 'ISS': 'INFY'}
    return mapping.get(symbol, symbol)

def place_order(symbol, transaction_type):
    """Place a market order."""
    try:
        mapped_symbol = map_symbol(symbol)
        order_id = kite.place_order(
            variety=kite.VARIETY_REGULAR,
            exchange=kite.EXCHANGE_NSE,
            tradingsymbol=mapped_symbol,
            transaction_type=transaction_type,
            quantity=quantity,
            order_type=kite.ORDER_TYPE_MARKET,
            product=product_type
        )
        log_action(f"{transaction_type} order placed for {symbol}. Order ID: {order_id}")

        # Wait for execution
        time.sleep(0.5)
        order_history = kite.order_history(order_id)
        for record in order_history:
            if record['status'] == 'COMPLETE':
                price = record['average_price']
                exec_time = datetime.now()
                log_price(exec_time, price, transaction_type)
                break
        return order_id
    except Exception as e:
        log_action(f"Order placement failed: {e}")
        return None

def square_off_position(symbol, side):
    """Square off an existing position."""
    reverse_side = 'SELL' if side == 'BUY' else 'BUY'
    order_id = place_order(symbol, reverse_side)
    if order_id:
        log_action(f"Position squared off for {symbol} via {reverse_side}")
    return order_id

def fetch_resampled_ohlc(symbol):
    """Fetch and resample OHLC and Volume for a symbol."""
    conn = sqlite3.connect(db_path)
    vol_col = f"{symbol}V"
    query = f"SELECT datestamp, `{symbol}`, `{vol_col}` FROM testTable"
    df = pd.read_sql_query(query, conn)
    conn.close()
    df['datestamp'] = pd.to_datetime(df['datestamp'])
    df.set_index('datestamp', inplace=True)
    ohlc = df[symbol].resample('1min').ohlc().dropna()
    df['volume_diff'] = df[vol_col].diff().clip(lower=0)
    vol = df['volume_diff'].resample('1min').sum()
    vol = vol.loc[ohlc.index]
    ohlc['volume'] = vol
    return ohlc

def detect_signal(ohlc):
    """Detect EMA crossover signal on last 2 completed candles."""
    if len(ohlc) < 16:
        return None
    ohlc['EMA5'] = ohlc['close'].ewm(span=5, adjust=False).mean()
    ohlc['EMA15'] = ohlc['close'].ewm(span=15, adjust=False).mean()
    prev_ema5 = ohlc['EMA5'].iloc[-3]
    prev_ema15 = ohlc['EMA15'].iloc[-3]
    curr_ema5 = ohlc['EMA5'].iloc[-2]
    curr_ema15 = ohlc['EMA15'].iloc[-2]
    if prev_ema5 < prev_ema15 and curr_ema5 > curr_ema15:
        return 'BUY'
    elif prev_ema5 > prev_ema15 and curr_ema5 < curr_ema15:
        return 'SELL'
    else:
        return None

# ====== MAIN LOOP ======
while True:
    try:
        # Check current open net positions
        positions = kite.positions()['net']
        net_positions = {p['tradingsymbol']: p['quantity'] for p in positions}

        live_trade = get_live_trade()

        if live_trade:
            symbol = live_trade['symbol']
            side = live_trade['side']
            mapped_symbol = map_symbol(symbol)

            open_qty = net_positions.get(mapped_symbol, 0)
            
            if open_qty != 0:
                ohlc = fetch_resampled_ohlc(symbol)
                signal = detect_signal(ohlc)                

                if signal and signal != side:
                    order_id = square_off_position(symbol, side)
                    if order_id:
                        remove_live_trade()
            else:
                remove_live_trade()

        else:
            if not is_order_pending() and all(qty == 0 for qty in net_positions.values()):
                for symbol in symbols:
                    ohlc = fetch_resampled_ohlc(symbol)
                    signal = detect_signal(ohlc)
                    if signal:
                        order_id = place_order(symbol, signal)
                        if order_id:
                            log_trade_taken(symbol, signal)
                        break  # one trade at a time

        time.sleep(poll_interval)

    except Exception as e:
        log_action(f"Main loop error: {e}")
        time.sleep(poll_interval)
